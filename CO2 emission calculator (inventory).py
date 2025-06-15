import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Step 1: Load the Excel file and read its contents
df = pd.read_excel("file8 (1).xlsx", sheet_name="Inventory", header=None)

# Step 2: Select only columns A and D from the sheet
filtered_df = df[[0, 3]]
filtered_df.columns = ["Component", "Total kg CO2 eq"]
filtered_df = filtered_df.dropna(how='all')

# Step 3: Save the filtered data into a new Excel file without applying any formatting yet
output_file = "columns_A_D_CO2_output.xlsx"
filtered_df.to_excel(output_file, index=False)

# Step 4: Open the saved Excel file again and start adding formatting and a cover title
wb = load_workbook(output_file)
ws = wb.active

# Add a blank row at the top of the sheet for the title
ws.insert_rows(1)

# Write a main title in cell A1
cover_title = "Lifecycle Stage, Components & Processes"
ws["A1"] = cover_title

# Overwrite the previous title by setting a new title in cell A2
cover_title = "Material Production"
ws["A2"] = cover_title

# Merge cells A1 and B1 to center the cover title across two columns
ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

# Apply font style, fill color, and alignment to the title cell
cover_fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")  # light blue
cover_font = Font(bold=True, size=14)
cover_alignment = Alignment(horizontal="center", vertical="center")

cell = ws["A1"]
cell.fill = cover_fill
cell.font = cover_font
cell.alignment = cover_alignment

# Format the second row (the header) with bold text and background color
header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
bold_font = Font(bold=True)

for cell in ws[2]:  # Second row is the header now
    cell.fill = header_fill
    cell.font = bold_font

# Make the title row taller so the title looks better
ws.row_dimensions[1].height = 30

# Step 5: Save all the formatting changes to the Excel file
wb.save(output_file)

print("✅ Excel file with cover title and formatted header created: 'columns_A_D_CO2_output.xlsx'")
