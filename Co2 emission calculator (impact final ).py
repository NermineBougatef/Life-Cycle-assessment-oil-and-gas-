import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Load the Excel file
file_path = "file8 (1).xlsx"
xls = pd.ExcelFile(file_path)
df = xls.parse("Impacts")

# Use the second row as headers
df.columns = df.iloc[1]
df = df.drop([0, 1]).reset_index(drop=True)

# Remove empty columns
df = df.dropna(axis=1, how='all')

# Select the useful columns
df_selected = df.iloc[:, [0, 1, 2, 3, 4]].copy()
co2_column = [col for col in df.columns if "CO2" in str(col)][0]
df_selected[co2_column] = df[co2_column].values

# Rename the columns
df_selected.columns = [
    'Component',
    'Design Input',
    'Design Unit',
    'Functional Unit',
    'Functional Input',
    'kg CO2 eq'
]

# Remove completely empty rows
df_selected = df_selected.dropna(how='all')

# Export to Excel
output_file = "extracted_component_co2.xlsx"
df_selected.to_excel(output_file, index=False)

# Reload the file for formatting
wb = load_workbook(output_file)
ws = wb.active

# Insert two rows at the top
ws.insert_rows(1, amount=2)

# Merge header groups
ws.merge_cells("A1:E1")
ws["A1"] = "Lifecycle Stage, Components & Processes"

ws.merge_cells("F1:F1")
ws["F1"] = "Impact Categories"

ws.merge_cells("A2:E2")
ws["A2"] = "Material Production"

ws.merge_cells("F2:F2")
ws["F2"] = "Total Impact"

# Define styles
header_font = Font(bold=True)
header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
center_align = Alignment(horizontal="center", vertical="center")

# Apply style to rows 1 and 2 (merged headers)
for row in [1, 2]:
    for cell in ws[row]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

# Apply style to the actual header row (row 3)
header_row = 3
for cell in ws[header_row]:
    if cell.value:  # Don't style empty cells
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

# Auto-adjust column widths
for i, col in enumerate(ws.columns, 1):
    max_length = 0
    col_letter = get_column_letter(i)
    for cell in col:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass
    ws.column_dimensions[col_letter].width = max_length + 2

# Save
wb.save(output_file)
print("✅ File 'extracted_component_co2.xlsx' generated with hierarchical structure and styled headers.")
